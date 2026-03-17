from flask import Blueprint, render_template, request, jsonify, send_file
from .extensions import db
from .models import Employee
from .payroll import PayrollCalculator
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. Сначала создаем Blueprint (Важно!)
main_bp = Blueprint('main', __name__)

# 2. Главная страница
@main_bp.route('/')
def index():
    history = Employee.query.order_by(Employee.id.desc()).limit(10).all()
    return render_template('index.html', history=history)

# 3. API: расчет зарплаты
@main_bp.route('/api/payroll', methods=['POST'])
def payroll():
    try:
        data = request.get_json(force=True)
        gross = float(data.get('gross_salary', 0))
        children = int(data.get('children', 0))
        age = int(data.get('age', 0))

        calc = PayrollCalculator(gross, children, age)
        result = calc.calculate_net()

        new_calc = Employee(
            gross_salary=gross,
            children=children,
            age=age,
            tax=result['tax'],
            net=result['net']
        )
        db.session.add(new_calc)
        db.session.commit()
        return jsonify(result)
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# 4. API: экспорт в Excel (Продвинутая версия со стилями и формулами)
@main_bp.route('/api/export-excel')
def export_excel():
    try:
        results = Employee.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll 2026"

        # Настройка стилей
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="005A9C", end_color="005A9C", fill_type="solid")
        center_aligned = Alignment(horizontal='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Заголовки
        headers = ['ID', 'Gross Salary', 'Age', 'Tax', 'Net Salary', 'Date']
        ws.append(headers)

        # Стилизуем заголовки
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
            cell.border = thin_border

        # Заполняем данными
        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.id)
            ws.cell(row=i, column=2, value=float(r.gross_salary))
            ws.cell(row=i, column=3, value=r.age)
            ws.cell(row=i, column=4, value=float(r.tax))
            
            # ФОРМУЛА EXCEL: Net = Gross - Tax
            # Колонка B - Gross, D - Tax, E - Net
            ws.cell(row=i, column=5, value=f"=B{i}-D{i}") 
            
            ws.cell(row=i, column=6, value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")

            # Форматирование ячеек (Границы и €)
            for col in range(1, 7):
                cell = ws.cell(row=i, column=col)
                cell.border = thin_border
                if col in [2, 4, 5]: # Денежные колонки
                    cell.number_format = '#,##0.00 €'

        # Авто-подбор ширины колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 3

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="payroll_pro_2026.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500